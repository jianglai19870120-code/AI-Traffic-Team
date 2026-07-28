#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover
    raise SystemExit("缺少依赖 openpyxl，无法读取标准 xlsx。") from exc

try:
    import xlrd
except ImportError:  # pragma: no cover
    xlrd = None

def bootstrap_brand_footer() -> None:
    candidate_roots: List[Path] = []

    if "--root" in sys.argv:
        try:
            root_arg = sys.argv[sys.argv.index("--root") + 1]
            candidate_roots.append(Path(root_arg).resolve())
        except (IndexError, OSError):
            pass

    cwd = Path.cwd().resolve()
    candidate_roots.extend([cwd, *cwd.parents])

    script_path = Path(__file__).resolve()
    candidate_roots.extend(script_path.parents)

    seen: set[str] = set()
    for root in candidate_roots:
        key = os.fspath(root)
        if key in seen:
            continue
        seen.add(key)
        tools_dir = root / "tools"
        if (tools_dir / "brand_footer.py").exists():
            sys.path.insert(0, str(tools_dir))
            return


bootstrap_brand_footer()

try:
    from brand_footer import append_brand_footer
except ImportError:
    def append_brand_footer(text: str) -> str:
        return text


CATEGORIES = {
    "科学创业": "01_科学创业选题表.md",
    "能力成长": "02_能力成长选题表.md",
    "赚钱财富": "03_赚钱财富选题表.md",
    "个人IP": "04_个人IP选题表.md",
    "AI科技": "05_AI科技选题表.md",
    "其他类型": "99_其他类型选题表.md",
}

TABLE_HEADER = ["选题", "主题分类", "博主名", "点赞数", "链接", "发布时间", "选中选题", "选中开头"]
LEGACY_TABLE_HEADER = ["选题", "主题分类", "博主名", "点赞数", "链接", "发布时间", "是否选用"]
MANUAL_TABLE_HEADER = ["序号", "文案结构", "选题", "链接", "正文时间", "正文状态", "成稿时间", "成稿状态", "视觉时间", "视觉状态"]
OPENING_SELECTION_HEADER = ["序号", "选题", "链接", "状态", "备注"]
TOPIC_TABLE_GLOB = "*选题表.md"
HAND_INPUT_FILENAME = "00_手动输入选题表.md"
VALID_SELECTION_VALUES = {"", "是", "否"}
XLS_SIGNATURE = bytes.fromhex("D0CF11E0A1B11AE1")
XLSX_SIGNATURE = b"PK"
SOFT_TOPIC_LIMIT = 34
LONG_COMPLETE_LIMIT = 56
BAD_ENDING_RE = re.compile(r"(一|二|三|四|五|六|七|八|九|十|第|首先|其次|然后|因为|但是|而且|以及|如果|所以|例如|比如|其中|另外)$")
FRAGMENT_SUFFIX_RE = re.compile(
    r"(过后|之后|之间|的时候|阶段|故事|套路|系列|视角|原因|经验分享|想法|感觉|真相|秘诀|方向|自由|标准|阶段后)$"
)
OPENING_PUNCT = "([{【「『“'\""
CLOSING_PUNCT = ")]}】」』”'\""
EXPLANATION_STARTERS = [
    "AIDA代表",
    "很多人",
    "大多数人",
    "其实",
    "我觉得",
    "我前面",
    "我每次",
    "我经历的",
    "我做",
    "我用",
    "这里说的",
    "关键在于",
    "也就是",
    "区别只在于",
    "本质是",
    "记住",
    "相信我",
    "说实话",
    "如果你",
    "如果你的",
    "最好的",
    "真正的",
    "核心是",
    "如何发现",
    "把问题变成机会",
    "你的人生",
    "做知识博主能赚钱吗",
    "普通人如何赚钱",
    "用你的结果",
    "穷人用时间换钱",
    "你花在",
    "这条视频",
    "起点很低",
    "普通人",
    "短视频创作",
    "做短视频",
    "卖货的本质",
    "新人做",
    "真正重要的是",
    "第一步",
    "第二步",
    "第三步",
    "一、",
    "二、",
    "三、",
    "1、",
    "2、",
    "3、",
]
WEAK_SHORT_TOPICS = {
    "不露脸",
    "你的时间",
    "一边学习",
    "先走心",
    "换张地图",
}
GENERIC_LEAD_PHRASES = [
    "1条视频讲清楚",
    "一条视频讲清楚",
    "写给小白",
    "这条视频讲清楚",
]
MEANINGFUL_TAIL_PATTERNS = [
    r"如何变现$",
    r"怎么赚钱$",
    r"的区别$",
    r"有什么区别$",
    r"是什么$",
    r"怎么做$",
    r"怎么选$",
    r"怎么用$",
    r"个启发$",
    r"个思考$",
    r"种模式$",
    r"种方法$",
    r"条路线$",
    r"干抖音$",
]


@dataclass
class TopicRow:
    topic: str
    primary_bucket: str
    display_category: str
    blogger: str
    likes: str
    link: str
    published_at: str
    selected_topic: str = ""
    selected_opening: str = ""

    def key(self) -> Tuple[str, str]:
        if self.link:
            return ("link", self.link)
        return ("topic", f"{self.blogger}::{self.topic}")

    def to_cells(self) -> List[str]:
        return [
            self.topic,
            self.display_category,
            self.blogger,
            self.likes,
            self.link,
            self.published_at,
            self.selected_topic,
            self.selected_opening,
        ]


def normalize_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value).strip()


def markdown_escape(value: str) -> str:
    value = normalize_cell(value)
    value = value.replace("\\", "\\\\").replace("|", "\\|")
    value = value.replace("\r\n", "<br>").replace("\n", "<br>").replace("\r", "<br>")
    return value


def split_markdown_row(line: str) -> List[str]:
    text = line.strip()
    if text.startswith("|"):
        text = text[1:]
    if text.endswith("|"):
        text = text[:-1]
    cells: List[str] = []
    current: List[str] = []
    escaped = False
    for char in text:
        if char == "\\" and not escaped:
            escaped = True
            current.append(char)
            continue
        if char == "|" and not escaped:
            cells.append("".join(current).strip().replace("\\|", "|"))
            current = []
        else:
            current.append(char)
        escaped = False
    cells.append("".join(current).strip().replace("\\|", "|"))
    return cells


def is_separator_row(cells: List[str]) -> bool:
    return all(re.fullmatch(r":?-{3,}:?", cell.strip()) for cell in cells if cell.strip())


def compact_text(raw: str) -> str:
    text = normalize_cell(raw)
    text = text.split("#", 1)[0]
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"(展开|收起|复制链接|DOU\+小助手)$", "", text, flags=re.I).strip()
    text = re.sub(r"([\s|｜/\-—]+抖音)$", "", text, flags=re.I).strip()
    return text


def split_sentences(text: str) -> List[str]:
    parts = re.split(r"(?<=[。！？?!])\s*|…{2,}|\.\.\.+", text)
    return [part.strip() for part in parts if part.strip()]


def normalize_topic_candidate(text: str) -> str:
    candidate = re.sub(r"\s+", " ", text).strip(" ，,；;：:。.!?！？、-…")
    return candidate.strip()


def dedupe_repeated_topic(text: str) -> str:
    normalized = normalize_topic_candidate(text)
    if not normalized:
        return normalized
    parts = [part.strip() for part in re.split(r"[。！？?!]", normalized) if part.strip()]
    if len(parts) >= 2 and parts[0] == parts[1]:
        return parts[0]
    if len(normalized) % 2 == 0:
        half = len(normalized) // 2
        left = normalized[:half].strip()
        right = normalized[half:].strip()
        if left and left == right:
            return left
    return normalized


def has_unmatched_opening(text: str) -> bool:
    stack: List[str] = []
    pairs = dict(zip(CLOSING_PUNCT, OPENING_PUNCT))
    symmetric_pairs = {'"', "'"}
    for char in text:
        if char in symmetric_pairs:
            if stack and stack[-1] == char:
                stack.pop()
            else:
                stack.append(char)
            continue
        if char in OPENING_PUNCT:
            stack.append(char)
        elif char in CLOSING_PUNCT:
            if stack and stack[-1] == pairs[char]:
                stack.pop()
    return bool(stack)


def ends_like_fragment(text: str) -> bool:
    if not text:
        return True
    if BAD_ENDING_RE.search(text):
        return True
    if FRAGMENT_SUFFIX_RE.search(text):
        return True
    if has_unmatched_opening(text):
        return True
    return bool(re.search(r"[：:（(「『“\"]$", text))


def looks_like_weak_short_topic(text: str) -> bool:
    if text in WEAK_SHORT_TOPICS:
        return True
    if text.endswith("？") or text.endswith("?"):
        return False
    if re.search(r"(做事|成事|资产|投资|成长|方法|路径|系统|模型|策略|财富|创业|赚钱|变现|销售|工作流|能力|认知|习惯)$", text):
        return False
    if len(text) <= 6 and not re.search(r"(如何|怎么|为什么|是否|会不会|能不能|是不是)", text):
        return True
    return False


def is_complete_topic(text: str) -> bool:
    candidate = normalize_topic_candidate(text)
    if not candidate:
        return False
    if ends_like_fragment(candidate):
        return False
    if looks_like_weak_short_topic(candidate):
        return False
    return True


def split_clauses(text: str) -> List[str]:
    return [part.strip() for part in re.split(r"(…{2,}|\.\.\.+|[，；;。！？?!])", text) if part and part.strip()]


def split_space_phrases(text: str) -> List[str]:
    return [part.strip() for part in text.split(" ") if part.strip()]


def looks_like_hook_phrase(text: str) -> bool:
    if not text:
        return False
    candidate = normalize_topic_candidate(text)
    if len(candidate) < 5 or len(candidate) > 28:
        return False
    if ends_like_fragment(candidate):
        return False
    return bool(re.search(r"(怎么|如何|为什么|能不能|有没有|什么时候|多少|区别|秘诀|方法|收入来源|赚钱|买房|起号|翻盘|变现|自由|启发|指南|攻略|真相|认知|策略)", candidate))


def trim_label_tail(text: str) -> str:
    raw_candidate = re.sub(r"\s+", " ", str(text)).strip()
    candidate = normalize_topic_candidate(raw_candidate)
    if not candidate:
        return ""
    explicit_followup = re.match(
        r"^(?P<head>.+?)\s+(?P<followup>(做|靠|用|学|买|选).{0,14}(赚钱吗|怎么赚钱|能赚钱吗|值不值|靠不靠谱|有没有必要|好不好|行不行))$",
        candidate,
    )
    if explicit_followup:
        head = normalize_topic_candidate(explicit_followup.group("head"))
        if len(head) >= 6 and is_complete_topic(head):
            return head
    raw_phrases = [part.strip() for part in raw_candidate.split(" ") if part.strip()]
    phrases = [normalize_topic_candidate(part) for part in raw_phrases]
    if len(phrases) >= 2:
        first = normalize_topic_candidate(phrases[0])
        raw_remainder = " ".join(raw_phrases[1:])
        remainder = normalize_topic_candidate(raw_remainder)
        if first and remainder:
            if ("？" in raw_remainder or "?" in raw_remainder) and len(first) >= 6:
                return first
            if looks_like_hook_phrase(remainder) and len(first) >= 6:
                return first
            if remainder.startswith(("你", "我", "如果", "当", "因为", "这是", "也就是", "普通人")) and len(first) >= 6:
                return first
    return candidate


def extract_structured_head(text: str) -> str:
    candidate = normalize_topic_candidate(text)
    if not candidate:
        return ""

    generic_lead = re.match(r"^(?P<lead>[^：:]{2,16})[：:\s]+(?P<body>.+)$", candidate)
    if generic_lead:
        lead = normalize_topic_candidate(generic_lead.group("lead"))
        body = normalize_topic_candidate(generic_lead.group("body"))
        if lead in GENERIC_LEAD_PHRASES and body and is_complete_topic(body):
            return body

    numbered_with_followup = re.match(
        r"^(?P<prefix>.+?)\s+[一二三四五六七八九十]+、(?P<section>[^。！？?!]{4,40}?)\s+(?P<followup>[^。！？?!]{4,24})$",
        candidate,
    )
    if numbered_with_followup:
        prefix = normalize_topic_candidate(numbered_with_followup.group("prefix"))
        section = normalize_topic_candidate(numbered_with_followup.group("section"))
        followup = normalize_topic_candidate(numbered_with_followup.group("followup"))
        if prefix and section and looks_like_hook_phrase(followup):
            combined = normalize_topic_candidate(f"{prefix}：{section}")
            if is_complete_topic(combined):
                return combined

    numbered = re.match(r"^(?P<prefix>.+?)\s+[一二三四五六七八九十]+、(?P<section>[^。！？?!]{4,30})", candidate)
    if numbered:
        prefix = normalize_topic_candidate(numbered.group("prefix"))
        section = trim_label_tail(numbered.group("section"))
        if prefix and section:
            combined = normalize_topic_candidate(f"{prefix}：{section}")
            if is_complete_topic(combined):
                return combined

    decimal_numbered = re.match(r"^(?P<prefix>.+?)\s+0\.(?P<section>[^。！？?!]{4,30})", candidate)
    if decimal_numbered:
        prefix = normalize_topic_candidate(decimal_numbered.group("prefix"))
        section = trim_label_tail(decimal_numbered.group("section"))
        if prefix and section:
            combined = normalize_topic_candidate(f"{prefix}：{section}")
            if is_complete_topic(combined):
                return combined

    lead_with_list = re.match(r"^(?P<intro>.+?)\s+(?P<body>[^。！？?!]{6,32}[：:])\s*[0-9一二三四五六七八九十]+[、.]", candidate)
    if lead_with_list:
        body = normalize_topic_candidate(lead_with_list.group("body").rstrip("：:"))
        if is_complete_topic(body):
            return body

    prefix_with_numbered_list = re.match(
        r"^(?P<head>[^。！？?!]{6,34}?)\s+[0-9一二三四五六七八九十]+[、:：]",
        candidate,
    )
    if prefix_with_numbered_list:
        head = normalize_topic_candidate(prefix_with_numbered_list.group("head"))
        if is_complete_topic(head):
            return head

    phrases = split_space_phrases(candidate)
    if len(phrases) >= 2:
        first = normalize_topic_candidate(phrases[0])
        second = normalize_topic_candidate(" ".join(phrases[1:]))
        if first in GENERIC_LEAD_PHRASES and second and is_complete_topic(second):
            return second
        if looks_like_hook_phrase(first) and second.startswith(("如果", "你", "普通人", "一个人", "我", "这", "做", "想", "当")):
            return first
        if first and second and re.match(r"^[0-9一二三四五六七八九十]+[、.]", second):
            return first

    return candidate


def first_sentence(text: str) -> str:
    sentences = split_sentences(text)
    return sentences[0] if sentences else text.strip()


def truncate_to_boundary(text: str, limit: int) -> str:
    if len(text) <= limit:
        return normalize_topic_candidate(text)
    boundary = max(
        text.rfind("，", 0, limit + 1),
        text.rfind("；", 0, limit + 1),
        text.rfind("。", 0, limit + 1),
        text.rfind(" ", 0, limit + 1),
        text.rfind("：", 0, limit + 1),
        text.rfind("、", 0, limit + 1),
    )
    if boundary >= 8:
        return normalize_topic_candidate(text[:boundary])
    return normalize_topic_candidate(text[:limit])


def trim_explanatory_tail(text: str) -> str:
    candidate = extract_structured_head(dedupe_repeated_topic(text))
    if not candidate:
        return ""
    space_parts = candidate.split(" ", 1)
    if len(space_parts) == 2:
        first = normalize_topic_candidate(space_parts[0])
        second = normalize_topic_candidate(space_parts[1])
        if first.endswith("就是") and second:
            merged = normalize_topic_candidate(f"{first} {second}")
            if len(merged) <= LONG_COMPLETE_LIMIT and is_complete_topic(merged):
                return merged
        if (
            4 <= len(first) <= 30
            and is_complete_topic(first)
            and second.startswith(("当", "如果", "人会", "我一直", "你周围", "投资", "现在", "\"", "“", "‘"))
        ):
            return first
    sentence_parts = re.split(r"[。！？?!]\s+", candidate, maxsplit=1)
    if len(sentence_parts) == 2:
        head = normalize_topic_candidate(sentence_parts[0])
        if 4 <= len(head) <= 30 and is_complete_topic(head):
            return head
    sentence_head = re.match(r"^(?P<head>[^。！？?!]{4,30}[。！？?!])\s+(?P<tail>.+)$", candidate)
    if sentence_head:
        head = normalize_topic_candidate(sentence_head.group("head"))
        if is_complete_topic(head):
            return head
    starter_head = re.match(
        r"^(?P<head>[^。！？?!]{4,30})\s+(?P<tail>(当|如果|人会|我一直|你周围|投资|现在|\"|“|‘).+)$",
        candidate,
    )
    if starter_head:
        head = normalize_topic_candidate(starter_head.group("head"))
        if is_complete_topic(head):
            return head
    judgement_pair = re.match(r"^(?P<head>[^。！？?!]{4,24}就是)\s+(?P<tail>[^。！？?!]{2,24})$", candidate)
    if judgement_pair:
        merged = normalize_topic_candidate(f"{judgement_pair.group('head')} {judgement_pair.group('tail')}")
        if is_complete_topic(merged):
            return merged
    if any(re.search(pattern, candidate, flags=re.I) for pattern in MEANINGFUL_TAIL_PATTERNS):
        return candidate
    if len(candidate) <= LONG_COMPLETE_LIMIT and is_complete_topic(candidate):
        return candidate
    lead_phrases = split_space_phrases(candidate)
    if len(lead_phrases) >= 2:
        first = normalize_topic_candidate(lead_phrases[0])
        second = normalize_topic_candidate(lead_phrases[1])
        if first.endswith("就是"):
            return candidate
        if (
            4 <= len(first) <= 28
            and is_complete_topic(first)
            and not any(re.search(pattern, second, flags=re.I) for pattern in MEANINGFUL_TAIL_PATTERNS)
            and second.startswith(("很多", "最好的", "普通人", "借", "学会", "把", "如何", "核心", "与", "就是", "真正", "关键", "第一步", "第二步", "第三步", "人会", "你", "我", "我们", "现在", "如果", "当", "不是", "想", "\"", "“", "‘"))
        ):
            return first
    starter_head = re.match(
        r"^(?P<head>.+?)\s+(?P<tail>(最好的|真正的|核心是|如何发现|把问题变成机会|你的人生|做知识博主能赚钱吗|普通人如何赚钱|用你的结果|穷人用时间换钱).+)$",
        candidate,
    )
    if starter_head:
        head = normalize_topic_candidate(starter_head.group("head"))
        if len(head) >= 6 and is_complete_topic(head):
            return head
    numbered_followup_head = re.match(
        r"^(?P<prefix>.+?)\s+[一二三四五六七八九十]+、(?P<section>[^。！？?!]{4,40}?)\s+(?P<followup>(做|靠|用|学|买|选).{0,14}(赚钱吗|怎么赚钱|能赚钱吗|值不值|靠不靠谱|有没有必要|好不好|行不行))(?:[。！？?!]|$)",
        candidate,
    )
    if numbered_followup_head:
        prefix = normalize_topic_candidate(numbered_followup_head.group("prefix"))
        section = normalize_topic_candidate(numbered_followup_head.group("section"))
        if prefix and section:
            combined = normalize_topic_candidate(f"{prefix}：{section}")
            if is_complete_topic(combined):
                return combined
    question_cut = max(candidate.find("？"), candidate.find("?"))
    if question_cut >= 8 and question_cut < len(candidate) - 1:
        head = normalize_topic_candidate(candidate[: question_cut + 1])
        if is_complete_topic(head):
            return head
    for starter in EXPLANATION_STARTERS:
        for token in (f" {starter}", f"，{starter}", f"：{starter}", f"。{starter}"):
            position = candidate.find(token)
            if position >= 8:
                head = normalize_topic_candidate(candidate[:position])
                if is_complete_topic(head):
                    return head
    return candidate


def compress_if_overlong(text: str) -> Tuple[str, bool, str]:
    normalized = dedupe_repeated_topic(text)
    if not normalized:
        return "", False, ""
    if len(normalized) <= SOFT_TOPIC_LIMIT and is_complete_topic(normalized):
        return normalized, False, ""
    if len(normalized) <= LONG_COMPLETE_LIMIT and is_complete_topic(normalized):
        return normalized, False, ""

    sentences = split_sentences(normalized)
    sentence = trim_explanatory_tail(first_sentence(normalized))
    if sentence != normalized and is_complete_topic(sentence) and len(sentence) <= LONG_COMPLETE_LIMIT:
        return sentence, True, "多句内容，只保留首个完整主张句"

    clauses = split_clauses(sentence)
    candidate = ""
    for piece in clauses:
        merged = normalize_topic_candidate(candidate + piece)
        if not merged:
            continue
        candidate = merged
        if len(candidate) <= SOFT_TOPIC_LIMIT and is_complete_topic(candidate):
            return candidate, True, "压缩口播解释串，保留最小完整句"

    phrases = split_space_phrases(sentence)
    if len(phrases) > 1:
        candidate = ""
        for part in phrases:
            merged = normalize_topic_candidate((candidate + " " + part).strip())
            if merged:
                candidate = merged
            if len(candidate) <= LONG_COMPLETE_LIMIT and is_complete_topic(candidate):
                trimmed = trim_explanatory_tail(candidate)
                return trimmed, True, "压缩空格串联解释，保留首个完整主张句"

    if is_complete_topic(sentence) and len(sentence) <= LONG_COMPLETE_LIMIT:
        return trim_explanatory_tail(sentence), True, "去掉后续展开，只保留首个完整主张句"

    running = ""
    best = sentence if sentence else normalized
    best_reason = "标题过长，保留首个完整分句"
    for piece in clauses:
        merged = normalize_topic_candidate(running + piece)
        if not merged:
            continue
        running = merged
        if len(running) <= LONG_COMPLETE_LIMIT and is_complete_topic(running):
            best = running
            best_reason = "标题过长，压缩为完整分句"

    if is_complete_topic(best):
        return trim_explanatory_tail(best), True, best_reason

    fallback = normalize_topic_candidate(sentence or normalized)
    bounded = truncate_to_boundary(fallback, LONG_COMPLETE_LIMIT)
    if bounded and is_complete_topic(bounded):
        return trim_explanatory_tail(bounded), True, "标题过长，按边界截到可用主张句"
    if fallback and not ends_like_fragment(fallback):
        return fallback, True, "标题过长，回退到最短可用主张句"

    return normalized, False, ""


def extract_primary_statement(text: str) -> Tuple[str, bool, str]:
    normalized = dedupe_repeated_topic(text)
    if not normalized:
        return "", False, ""
    trimmed = trim_explanatory_tail(normalized)
    if trimmed != normalized and is_complete_topic(trimmed) and len(trimmed) <= LONG_COMPLETE_LIMIT:
        return trimmed, True, "压缩解释串，保留最小完整主张句"
    if is_complete_topic(normalized) and len(normalized) <= LONG_COMPLETE_LIMIT:
        return normalized, False, ""
    return compress_if_overlong(normalized)


def clean_topic(raw: str) -> Tuple[str, bool, str, str]:
    text = compact_text(raw)
    if not text:
        return "", False, "", ""
    topic, shortened, reason = extract_primary_statement(text)
    return topic, shortened, text, reason


def score_keywords(text: str, keywords: Iterable[str]) -> int:
    lower_text = text.lower()
    score = 0
    for kw in keywords:
        normalized_kw = kw.lower()
        if re.search(r"[a-z0-9]", normalized_kw):
            pattern = re.compile(rf"(?<![a-z0-9]){re.escape(normalized_kw)}(?![a-z0-9])", re.I)
            if pattern.search(lower_text):
                score += 1
        elif normalized_kw in lower_text:
            score += 1
    return score


def infer_suspected_category(topic: str) -> Tuple[str, str]:
    lower_text = topic.lower()

    ai_keywords = [
        "ai", "agent", "gpt", "claude", "karpathy", "deepseek", "lex fridman", "prompt",
        "自动化", "工作流", "模型", "网站", "数据中心", "软件", "gui", "cli", "编码", "编程",
    ]
    ip_keywords = [
        "自媒体", "账号", "涨粉", "起号", "卖课", "课程", "知识博主", "赛道", "大主播",
        "内容", "女粉", "流量", "私域", "个人ip", "个人品牌", "选题", "文案",
    ]
    growth_keywords = [
        "成事", "做事", "做成", "执行力", "专注", "时间管理", "上班", "打工", "不上班",
        "空窗期", "面试", "方向", "喜欢", "感兴趣", "翻盘", "找工作", "社交", "人生",
        "个人提升", "成长", "习惯", "解决问题", "信息源", "生命周期",
    ]
    wealth_keywords = [
        "赚钱", "收入", "副业", "搞钱", "财富", "投资", "理财", "保险", "买房", "租房",
        "房租", "月供", "中介", "谈价", "房子", "房价", "金价", "关税", "贸易战",
        "经济", "医保", "翻倍", "变现",
    ]
    startup_keywords = [
        "创业", "商业", "生意", "利润", "客户", "产品", "增长", "成交", "销售", "过滤客户",
        "高端客户", "经营", "团队", "公司",
    ]

    if any(keyword in lower_text for keyword in ai_keywords):
        return "AI科技", "含 AI / 模型 / 自动化 / 科技产品 等语义"
    if any(keyword in lower_text for keyword in ip_keywords):
        return "个人IP", "含 自媒体 / 账号 / 涨粉 / 卖课 / 赛道 等语义"
    if any(keyword in lower_text for keyword in growth_keywords):
        return "能力成长", "含 成事 / 做事 / 职场处境 / 个人提升 等语义"
    if any(keyword in lower_text for keyword in wealth_keywords):
        return "赚钱财富", "含 收入 / 财富 / 房产 / 宏观经济 / 变现 等语义"
    if any(keyword in lower_text for keyword in startup_keywords):
        return "科学创业", "含 创业 / 生意 / 利润 / 客户 / 产品经营 等语义"
    return "待细化主题", "当前仍未命中稳定规则，需要继续补充通用分类口径"


def infer_other_display_category(topic: str, blogger: str) -> Tuple[str, str]:
    lower_text = f"{blogger} {topic}".lower()

    if any(keyword in lower_text for keyword in ["room tour", "四房两卫", "串串房", "居住展示", "入住", "一镜到底"]):
        return "居住展示记录", "主题聚焦居住空间、入住展示或看房后体验"

    if any(keyword in lower_text for keyword in ["城市", "base", "成都", "北京", "换城市", "定居", "生活方式"]):
        return "城市迁移与居住选择", "主题聚焦城市迁移、定居选择或生活方式切换"

    if any(keyword in lower_text for keyword in ["狗狗", "宠物", "大理", "玉龙雪山", "带狗旅行"]):
        return "宠物旅行", "主题聚焦宠物陪伴下的旅行或出行经验"

    if any(keyword in lower_text for keyword in ["采访", "对话", "播客截选", "访谈视频", "创始人对话", "截选了一小段"]):
        return "人物访谈片段", "主题聚焦人物访谈、播客截选或对话内容"

    if any(keyword in lower_text for keyword in ["年会", "线下", "厦门年会", "长沙线下", "活动现场"]):
        return "线下活动记录", "主题聚焦年会、线下见面或活动现场记录"

    if any(keyword in lower_text for keyword in ["我上央视了", "近况通报", "上央视"]):
        return "个人事件播报", "主题聚焦个人事件、近况或阶段性播报"

    if any(keyword in lower_text for keyword in ["信息密度", "刷视频状态", "内容观察", "编导在线答疑", "转发", "小作文"]):
        return "内容方法观察", "主题聚焦内容表达、平台观看状态或内容方法观察"

    if any(keyword in lower_text for keyword in ["科技现象", "前沿", "agi", "华为", "小米", "美国制造业", "豆包"]):
        return "科技现象观察", "主题聚焦技术前沿、科技公司或科技现象观察"

    return "抽象认知观点", "主题是泛认知、抽象判断或高概念观点，暂不稳定落入正式五类"


def classify_topic(topic: str) -> Tuple[str, str]:
    text = topic
    topic_text = topic

    ai_keywords = [
        "ai", "人工智能", "大模型", "模型", "agent", "智能体", "自动化", "编程", "代码",
        "claude", "gemini", "gpt", "openai", "cursor", "vibe coding", "芯片", "机器人",
        "工作流", "提示词", "token", "lex fridman", "karpathy", "deepseek",
    ]
    ip_keywords = [
        "自媒体", "短视频", "内容", "账号", "涨粉", "流量", "私域", "个人ip", "ip",
        "定位", "粉丝", "直播", "小红书", "抖音", "视频号", "公众号", "成交", "获客",
        "爆款", "文案", "选题", "剪辑", "起号", "卖课", "课程", "知识博主", "女粉",
        "内容产品", "个人品牌",
    ]
    money_keywords = [
        "赚钱", "收入", "副业", "财富", "现金流", "变现", "赚到", "赚", "钱",
        "商业机会", "普通人", "搞钱", "财富自由", "咨询", "接单", "投资", "理财",
        "生意", "创业赚钱",
    ]
    growth_keywords = [
        "学习", "认知", "效率", "行动", "行动力", "决策", "表达", "习惯", "复盘",
        "自律", "成长", "能力", "职业", "思考", "心态", "选择", "拖延", "焦虑",
        "知识", "做事", "现卖现学", "闭环", "执行力", "专注", "专注力", "时间管理",
        "上班", "打工", "空窗期", "面试", "人生方向", "自由职业", "成事", "小事",
        "做成", "超级个体", "个人提升",
    ]
    startup_keywords = [
        "创业", "商业模式", "公司", "企业", "产品", "组织", "增长", "战略", "融资",
        "管理", "团队", "企业服务", "创新", "商业化", "老板", "ceo", "客户",
    ]

    topic_scores = {
        "AI科技": score_keywords(topic_text, ai_keywords),
        "个人IP": score_keywords(topic_text, ip_keywords),
        "赚钱财富": score_keywords(topic_text, money_keywords),
        "能力成长": score_keywords(topic_text, growth_keywords),
        "科学创业": score_keywords(topic_text, startup_keywords),
    }
    scores = {
        name: topic_scores[name]
        for name in topic_scores
    }

    lower_text = text.lower()

    # 主承诺边界：账号增长和内容获客优先；上班/打工/执行力等个人处境优先成长；AI 工具和应用优先 AI科技。
    if scores["个人IP"] >= 1 and any(k in lower_text for k in ["涨粉", "流量", "账号", "自媒体", "短视频", "私域", "获客", "爆款", "文案", "选题", "起号", "卖课", "课程", "知识博主", "女粉"]):
        return "个人IP", "主承诺围绕内容账号、流量或个人品牌"

    if scores["AI科技"] >= 1 and any(k in lower_text for k in ["ai", "人工智能", "大模型", "模型", "agent", "智能体", "自动化", "claude", "gpt", "gemini", "cursor", "vibe coding", "工作流", "token", "提示词"]):
        if scores["个人IP"] == 0 or "ai赚钱" in lower_text or "用ai" in lower_text:
            return "AI科技", "主承诺围绕 AI、工具、模型或自动化应用"

    if any(k in text for k in ["买房", "租房", "房租", "月供", "房价", "房子", "中介", "谈价", "医保", "保险", "金价", "关税", "贸易战", "通货收缩"]):
        return "赚钱财富", "主承诺围绕财富决策、房产或宏观经济"

    if any(k in text for k in ["赛道", "大主播", "知识博主", "卖课", "起号", "涨粉", "女粉", "自媒体怎么", "做内容", "内容团队"]):
        return "个人IP", "主承诺围绕做账号、做内容或个人IP经营"

    if any(k in text for k in ["成事", "做事", "做成", "上班", "打工", "不上班", "空窗期", "面试", "找工作", "社交指南", "人生", "方向", "执行力", "专注力", "时间管理", "经验产品化", "信息源", "解决问题", "感兴趣", "喜欢的事", "翻盘"]):
        return "能力成长", "主承诺围绕个人成长、做事能力或职场处境"

    if any(k in lower_text for k in ["karpathy", "deepseek", "lex fridman", "gavin baker", "网站", "数据中心", "软件", "gui", "cli"]) and scores["赚钱财富"] == 0:
        return "AI科技", "主承诺围绕 AI 观察、科技产品或技术范式"

    if any(k in text for k in ["利润", "高端客户", "过滤客户", "低端客户", "生意越小", "规模做小", "做利润"]):
        return "科学创业", "主承诺围绕利润、客户筛选或经营策略"

    if scores["能力成长"] >= 1 and any(k in text for k in ["执行力", "专注", "专注力", "时间管理", "上班", "打工", "空窗期", "面试", "人生方向", "自律", "拖延", "心态", "成事", "做事", "做成", "小事", "超级个体", "个人提升"]):
        if scores["个人IP"] == 0:
            return "能力成长", "主承诺围绕个人成长、职业处境或能力提升"

    if scores["赚钱财富"] >= 2 and any(k in text for k in ["赚钱", "副业", "收入", "赚到", "变现", "接单", "财富", "生意"]):
        return "赚钱财富", "主承诺围绕赚钱、收入或变现"

    if topic_scores["AI科技"] > 0 and scores["AI科技"] >= 2 and scores["个人IP"] == 0 and scores["能力成长"] == 0:
        return "AI科技", "主承诺围绕 AI、工具、模型或科技产品"

    if topic_scores["能力成长"] > 0 and scores["个人IP"] == 0:
        return "能力成长", "主承诺围绕能力、认知或自我成长"

    if scores["科学创业"] >= 2 and scores["赚钱财富"] == 0 and scores["个人IP"] == 0:
        return "科学创业", "主承诺围绕创业、公司经营或商业化"

    if scores["能力成长"] >= 2 and scores["赚钱财富"] == 0:
        return "能力成长", "主承诺围绕能力、认知或自我成长"

    if scores["个人IP"] >= 1 and scores["赚钱财富"] >= 1:
        if any(k in lower_text for k in ["自媒体", "账号", "涨粉", "起号", "卖课", "课程", "知识博主", "女粉"]):
            return "个人IP", "同时涉及变现与账号，但主承诺更偏个人IP"

    if scores["能力成长"] >= 1 and scores["赚钱财富"] >= 1:
        if any(k in text for k in ["上班", "打工", "空窗期", "面试", "专注力", "时间管理", "执行力", "成事", "做事", "做成", "小事", "超级个体"]):
            return "能力成长", "同时涉及赚钱与成长，但主承诺更偏个人成长"

    best_category, best_score = max(scores.items(), key=lambda item: item[1])
    if best_score <= 0:
        return "其他类型", "信息不足或不属于当前五类"

    tied = [name for name, score in scores.items() if score == best_score]
    if len(tied) > 1:
        if "AI科技" in tied and topic_scores["AI科技"] > 0 and scores["赚钱财富"] == 0 and scores["个人IP"] == 0:
            return "AI科技", ""
        return "其他类型", "多类关键词冲突，需人工复核：" + "、".join(tied)

    return best_category, ""


def parse_markdown_table(path: Path, primary_bucket: str) -> Dict[Tuple[str, str], TopicRow]:
    rows: Dict[Tuple[str, str], TopicRow] = {}
    if not path.exists():
        return rows
    content = path.read_text(encoding="utf-8-sig")
    lines = [line for line in content.splitlines() if line.strip().startswith("|")]
    if len(lines) < 2:
        return rows
    header = split_markdown_row(lines[0])
    if header not in (TABLE_HEADER, LEGACY_TABLE_HEADER):
        return rows
    is_legacy = header == LEGACY_TABLE_HEADER
    for line in lines[1:]:
        cells = split_markdown_row(line)
        if is_separator_row(cells):
            continue
        if is_legacy and len(cells) != len(LEGACY_TABLE_HEADER):
            continue
        if not is_legacy and len(cells) != len(TABLE_HEADER):
            continue
        row = TopicRow(
            topic=cells[0],
            primary_bucket=primary_bucket,
            display_category=cells[1],
            blogger=cells[2],
            likes=cells[3],
            link=cells[4],
            published_at=cells[5],
            selected_topic=cells[6] if len(cells) >= 7 else "",
            selected_opening="" if is_legacy else cells[7],
        )
        rows[row.key()] = row
    return rows


def write_category_table(path: Path, category: str, rows: List[TopicRow]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        f"# {category}选题表",
        "",
        "| " + " | ".join(TABLE_HEADER) + " |",
        "|---|---|---|---:|---|---|---|---|",
    ]
    for row in rows:
        lines.append("| " + " | ".join(markdown_escape(v) for v in row.to_cells()) + " |")
    path.write_text(append_brand_footer("\n".join(lines)), encoding="utf-8")


def detect_excel_format(path: Path) -> str:
    with path.open("rb") as file_obj:
        head = file_obj.read(8)
    if head.startswith(XLSX_SIGNATURE):
        return "xlsx"
    if head.startswith(XLS_SIGNATURE):
        return "xls"
    return "unknown"


def read_xlsx(path: Path) -> Tuple[List[str], List[List[str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    headers = next(iterator, None)
    if not headers:
        return [], []
    normalized_headers = [normalize_cell(h) for h in headers]
    rows: List[List[str]] = []
    for row in iterator:
        rows.append([normalize_cell(v) for v in row])
    return normalized_headers, rows


def read_xls(path: Path) -> Tuple[List[str], List[List[str]]]:
    if xlrd is None:
        raise RuntimeError("缺少依赖 xlrd，无法读取老式 Excel。")
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    if sheet.nrows <= 0:
        return [], []
    headers = [normalize_cell(value) for value in sheet.row_values(0)]
    rows: List[List[str]] = []
    for row_index in range(1, sheet.nrows):
        rows.append([normalize_cell(value) for value in sheet.row_values(row_index)])
    return headers, rows


def read_excel_rows(path: Path) -> Tuple[str, List[str], List[List[str]]]:
    file_format = detect_excel_format(path)
    if file_format == "xlsx":
        headers, rows = read_xlsx(path)
        return file_format, headers, rows
    if file_format == "xls":
        headers, rows = read_xls(path)
        return file_format, headers, rows
    raise RuntimeError("无法识别的 Excel 文件格式")


def iter_input_files(input_dir: Path, blogger: Optional[str]) -> List[Path]:
    files = sorted({*input_dir.glob("*.xlsx"), *input_dir.glob("*.xls")})
    if blogger:
        files = [p for p in files if p.stem == blogger]
    return files


def classify_file(path: Path, audit: Dict[str, List[str]]) -> List[TopicRow]:
    blogger = path.stem
    try:
        file_format, headers, data_rows = read_excel_rows(path)
    except Exception as exc:
        audit["unreadable"].append(f"{path.name}：{type(exc).__name__} {exc}")
        return []

    if "视频信息" not in headers:
        audit["missing_required"].append(f"{path.name}：缺少 视频信息 列")
        return []

    header_index = {name: i for i, name in enumerate(headers)}
    result: List[TopicRow] = []
    for offset, cells in enumerate(data_rows, start=2):
        def get(name: str) -> str:
            i = header_index.get(name)
            if i is None or i >= len(cells):
                return ""
            return cells[i]

        raw_topic = get("视频信息")
        topic, shortened, classify_basis, shorten_reason = clean_topic(raw_topic)
        if not topic:
            audit["missing_topic"].append(f"{path.name} 第{offset}行：视频信息为空")
            continue

        if shortened:
            audit["long_topic"].append(
                f"{path.name} 第{offset}行：原始标题={classify_basis} -> 提炼后={topic}；原因={shorten_reason or '标题过长，保留完整主张句'}"
            )

        primary_bucket, _ = classify_topic(topic)
        display_category = primary_bucket
        if primary_bucket == "其他类型":
            display_category, display_reason = infer_other_display_category(topic, blogger)
            suspected_category, suspected_reason = infer_suspected_category(topic)
            suspected_label = "无" if suspected_category == "待细化主题" else suspected_category
            suspected_flag = "否" if suspected_label == "无" else "是"
            audit["other_reason"].append(
                f"{path.name} 第{offset}行：进入 99 缓冲区 -> 选题={topic}；当前具体命名={display_category}；是否疑似应归入五大类={suspected_flag}；疑似应归类={suspected_label}；原因：{display_reason if suspected_flag == '否' else suspected_reason}"
            )
        row = TopicRow(
            topic=topic,
            primary_bucket=primary_bucket,
            display_category=display_category,
            blogger=blogger,
            likes=get("点赞数"),
            link=get("链接"),
            published_at=get("发布时间"),
        )
        if not row.link:
            audit["missing_link"].append(f"{path.name} 第{offset}行：缺少链接，用 博主名+选题 去重")
        result.append(row)
    audit["readable"].append(f"{path.name}：按 {file_format} 读取 {len(data_rows)} 行，生成候选 {len(result)} 条")
    return result


def topic_table_paths(output_dir: Path) -> List[Path]:
    paths = []
    for path in sorted(output_dir.glob(TOPIC_TABLE_GLOB)):
        if path.name == HAND_INPUT_FILENAME:
            continue
        paths.append(path)
    return paths


def normalize_selection_value(value: str) -> str:
    text = normalize_cell(value)
    if text in VALID_SELECTION_VALUES:
        return text
    return text


def manual_selection_key(topic: str, link: str) -> Tuple[str, str]:
    normalized_link = normalize_cell(link)
    if normalized_link:
        return ("link", normalized_link)
    return ("topic", normalize_cell(topic))


def parse_manual_selection_table(path: Path) -> Tuple[List[str], Dict[Tuple[str, str], List[str]]]:
    rows: Dict[Tuple[str, str], List[str]] = {}
    if not path.exists():
        return MANUAL_TABLE_HEADER, rows
    content = path.read_text(encoding="utf-8-sig")
    table_lines = [line for line in content.splitlines() if line.strip().startswith("|")]
    if len(table_lines) < 2:
        return MANUAL_TABLE_HEADER, rows
    header = split_markdown_row(table_lines[0])
    legacy_header = ["序号", "文案结构", "选题", "正文时间", "正文状态", "成稿时间", "成稿状态", "视觉时间", "视觉状态"]
    if header != MANUAL_TABLE_HEADER and header != legacy_header:
        return MANUAL_TABLE_HEADER, rows
    for line in table_lines[1:]:
        cells = split_markdown_row(line)
        if is_separator_row(cells) or len(cells) != len(header):
            continue
        topic = cells[2].strip()
        if topic:
            if header == MANUAL_TABLE_HEADER:
                key = manual_selection_key(topic, cells[3].strip())
                rows[key] = cells
            else:
                upgraded = [cells[0], cells[1], cells[2], "", cells[3], cells[4], cells[5], cells[6], cells[7], cells[8]]
                key = manual_selection_key(topic, "")
                rows[key] = upgraded
    return MANUAL_TABLE_HEADER, rows


def write_manual_selection_table(path: Path, rows: List[List[str]]) -> None:
    lines = [
        "| " + " | ".join(MANUAL_TABLE_HEADER) + " |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for index, cells in enumerate(rows, start=1):
        current = list(cells)
        current[0] = str(index)
        lines.append("| " + " | ".join(markdown_escape(value) for value in current) + " |")
    path.write_text(append_brand_footer("\n".join(lines)), encoding="utf-8")


def parse_opening_selection_table(path: Path) -> Dict[Tuple[str, str], List[str]]:
    rows: Dict[Tuple[str, str], List[str]] = {}
    if not path.exists():
        return rows
    content = path.read_text(encoding="utf-8-sig")
    table_lines = [line for line in content.splitlines() if line.strip().startswith("|")]
    if len(table_lines) < 2:
        return rows
    header = split_markdown_row(table_lines[0])
    legacy_header = ["博主名", "视频信息", "链接", "状态", "备注"]
    if header != OPENING_SELECTION_HEADER and header != legacy_header:
        return rows
    for line in table_lines[1:]:
        cells = split_markdown_row(line)
        if is_separator_row(cells) or len(cells) != len(header):
            continue
        if header == OPENING_SELECTION_HEADER:
            key = opening_selection_key(cells[2], cells[1])
            rows[key] = cells
        else:
            upgraded = ["", cells[1], cells[2], cells[3], cells[4]]
            key = opening_selection_key(cells[2], cells[1])
            rows[key] = upgraded
    return rows


def opening_selection_key(link: str, topic: str) -> Tuple[str, str]:
    normalized_link = normalize_cell(link)
    if normalized_link:
        return ("link", normalized_link)
    return ("topic", normalize_cell(topic))


def write_opening_selection_table(path: Path, rows: List[List[str]]) -> None:
    lines = [
        "# 爆款开头选中清单",
        "",
        "|" + "|".join(OPENING_SELECTION_HEADER) + "|",
        "|-|-|-|-|-|",
    ]
    for index, cells in enumerate(rows, start=1):
        current = list(cells)
        current[0] = str(index)
        lines.append("|" + "|".join(markdown_escape(value) for value in current) + "|")
    path.write_text(append_brand_footer("\n".join(lines)), encoding="utf-8")


def build_manual_selection_row(topic_row: TopicRow, existing: Optional[List[str]]) -> List[str]:
    if existing:
        row = list(existing)
        row[2] = topic_row.topic
        row[3] = topic_row.link
        return row
    return ["", "", topic_row.topic, topic_row.link, "", "", "", "", "", ""]


def build_opening_selection_row(topic_row: TopicRow, existing: Optional[List[str]]) -> List[str]:
    if existing:
        row = list(existing)
        row[1] = topic_row.topic
        row[2] = topic_row.link
        row[3] = row[3] or "待拆解"
        row[4] = row[4] or "由选题表选中开头自动同步"
        return row
    return ["", topic_row.topic, topic_row.link, "待拆解", "由选题表选中开头自动同步"]


def sync_selection_outputs(output_dir: Path, audit: Dict[str, List[str]], totals: Dict[str, int]) -> None:
    manual_path = output_dir / "00_爆款选题选中清单.md"
    legacy_manual_path = output_dir / HAND_INPUT_FILENAME
    opening_path = output_dir.parent / "05_爆款开头库" / "00_爆款开头选中清单.md"

    _, existing_manual = parse_manual_selection_table(manual_path)
    if not existing_manual and legacy_manual_path.exists():
        _, existing_manual = parse_manual_selection_table(legacy_manual_path)
    existing_opening = parse_opening_selection_table(opening_path)

    selected_topics: List[TopicRow] = []
    selected_openings: List[TopicRow] = []
    deleted_from_sources = 0
    invalid_values = 0

    for table_path in topic_table_paths(output_dir):
        primary_bucket = next((name for name, filename in CATEGORIES.items() if filename == table_path.name), "其他类型")
        raw_rows = list(parse_markdown_table(table_path, primary_bucket).values())
        retained_rows: List[TopicRow] = []
        changed = False
        for row in raw_rows:
            row.selected_topic = normalize_selection_value(row.selected_topic)
            row.selected_opening = normalize_selection_value(row.selected_opening)
            for value_name, value in (("选中选题", row.selected_topic), ("选中开头", row.selected_opening)):
                if value not in VALID_SELECTION_VALUES:
                    audit["invalid_selection"].append(
                        f"{table_path.name}：选题={row.topic}；字段={value_name}；填值={value}；仅允许 是/否/留空"
                    )
                    invalid_values += 1
            if row.selected_topic == "否" and row.selected_opening == "否":
                deleted_from_sources += 1
                changed = True
                continue
            retained_rows.append(row)
            if row.selected_topic == "是":
                selected_topics.append(row)
            if row.selected_opening == "是":
                selected_openings.append(row)
        if changed:
            write_category_table(table_path, primary_bucket, retained_rows)

    manual_rows: List[List[str]] = []
    seen_manual_keys: set[Tuple[str, str]] = set()
    for row in selected_topics:
        key = manual_selection_key(row.topic, row.link)
        if key in seen_manual_keys:
            continue
        seen_manual_keys.add(key)
        manual_rows.append(build_manual_selection_row(row, existing_manual.get(key)))
    write_manual_selection_table(manual_path, manual_rows)
    if legacy_manual_path.exists():
        legacy_manual_path.unlink()

    opening_rows: List[List[str]] = []
    seen_opening_keys: set[Tuple[str, str]] = set()
    for row in selected_openings:
        key = opening_selection_key(row.link, row.topic)
        if key in seen_opening_keys:
            continue
        seen_opening_keys.add(key)
        opening_rows.append(build_opening_selection_row(row, existing_opening.get(key)))
    write_opening_selection_table(opening_path, opening_rows)

    totals["selected_topic_entries"] = len(manual_rows)
    totals["selected_opening_entries"] = len(opening_rows)
    totals["deleted_from_sources"] = deleted_from_sources
    totals["invalid_selection_inputs"] = invalid_values


def load_existing(output_dir: Path) -> Dict[str, Dict[Tuple[str, str], TopicRow]]:
    existing: Dict[str, Dict[Tuple[str, str], TopicRow]] = {}
    for category, filename in CATEGORIES.items():
        existing[category] = parse_markdown_table(output_dir / filename, category)
    return existing


def find_existing_category(existing: Dict[str, Dict[Tuple[str, str], TopicRow]], key: Tuple[str, str]) -> Optional[str]:
    for category, rows in existing.items():
        if key in rows:
            return category
    return None


def write_audit(audit_dir: Path, audit: Dict[str, List[str]], totals: Dict[str, int], output_dir: Path) -> Path:
    audit_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    path = audit_dir / f"{stamp}_爆款选题分类审核.md"
    lines = [
        "# 爆款选题分类审核",
        "",
        f"- 输出目录：{output_dir}",
        f"- 写入新选题：{totals['inserted']}",
        f"- 更新旧选题：{totals.get('updated', 0)}",
        f"- 跳过重复：{totals['duplicates']}",
        f"- 移动分类：{totals.get('moved', 0)}",
        f"- 本轮新进入其他类型：{totals['other']}",
        f"- 当前其他类型总量：{totals.get('other_total', totals['other'])}",
        f"- 正文入口同步条数：{totals.get('selected_topic_entries', 0)}",
        f"- 开头入口同步条数：{totals.get('selected_opening_entries', 0)}",
        f"- 双否删源条数：{totals.get('deleted_from_sources', 0)}",
        f"- 无效人工选择填值：{totals.get('invalid_selection_inputs', 0)}",
        "",
    ]
    sections = [
        ("可读取账号表", "readable"),
        ("不可读取账号表", "unreadable"),
        ("缺少必需字段", "missing_required"),
        ("缺少视频信息", "missing_topic"),
        ("缺少链接", "missing_link"),
        ("长选题提炼", "long_topic"),
        ("其他类型原因", "other_reason"),
        ("无效人工选择填值", "invalid_selection"),
    ]
    for title, key in sections:
        lines.extend([f"## {title}", ""])
        items = audit.get(key, [])
        if not items:
            lines.append("- 无")
        else:
            lines.extend(f"- {item}" for item in items)
        lines.append("")
    path.write_text(append_brand_footer("\n".join(lines)), encoding="utf-8")
    return path


def legacy_manual_table_state(output_dir: Path) -> Tuple[bool, Optional[float]]:
    manual = output_dir / HAND_INPUT_FILENAME
    if manual.exists():
        return True, manual.stat().st_mtime
    return False, None


def main() -> int:
    parser = argparse.ArgumentParser(description="从对标账号库分类生成爆款选题表")
    parser.add_argument("--root", default=None, help="工作区根目录，默认读取 AI_TRAFFIC_FACTORY_ROOT")
    parser.add_argument("--input-dir", default=None, help="账号表目录")
    parser.add_argument("--output-dir", default=None, help="爆款选题库输出目录")
    parser.add_argument("--audit-dir", default=None, help="审核报告目录")
    parser.add_argument("--blogger", default=None, help="只处理指定博主文件名，不含扩展名")
    args = parser.parse_args()

    script_path = Path(__file__).resolve()
    root_value = args.root or os.environ.get("AI_TRAFFIC_FACTORY_ROOT")
    root = Path(root_value).resolve() if root_value else script_path.parents[3]
    input_dir = Path(args.input_dir).resolve() if args.input_dir else root / "02_资产中心" / "03_对标账号库"
    output_dir = Path(args.output_dir).resolve() if args.output_dir else root / "02_资产中心" / "04_爆款选题库"
    audit_dir = Path(args.audit_dir).resolve() if args.audit_dir else root / "03_工作流中心" / "01_短视频主工作流" / "99_运行记录"

    if not input_dir.exists():
        raise SystemExit(f"输入目录不存在：{input_dir}")
    output_dir.mkdir(parents=True, exist_ok=True)

    legacy_manual_exists_before, legacy_manual_mtime_before = legacy_manual_table_state(output_dir)
    existing = load_existing(output_dir)
    audit: Dict[str, List[str]] = {
        "readable": [],
        "unreadable": [],
        "missing_required": [],
        "missing_topic": [],
        "missing_link": [],
        "long_topic": [],
        "other_reason": [],
        "invalid_selection": [],
    }
    totals = {"inserted": 0, "updated": 0, "duplicates": 0, "other": 0, "moved": 0}

    for file_path in iter_input_files(input_dir, args.blogger):
        for row in classify_file(file_path, audit):
            key = row.key()
            existing_category = find_existing_category(existing, key)
            if existing_category == row.primary_bucket:
                current = existing[row.primary_bucket].get(key)
                if current:
                    row.selected_topic = current.selected_topic
                    row.selected_opening = current.selected_opening
                if current and current.to_cells() == row.to_cells():
                    totals["duplicates"] += 1
                else:
                    existing[row.primary_bucket][key] = row
                    totals["updated"] += 1
                continue
            if existing_category and existing_category != row.primary_bucket:
                current = existing[existing_category].get(key)
                if current:
                    row.selected_topic = current.selected_topic
                    row.selected_opening = current.selected_opening
                existing[existing_category].pop(key, None)
                totals["moved"] += 1
            bucket = existing.setdefault(row.primary_bucket, {})
            bucket[key] = row
            if not existing_category:
                totals["inserted"] += 1
            if row.primary_bucket == "其他类型":
                totals["other"] += 1

    for category, filename in CATEGORIES.items():
        rows = sorted(existing.get(category, {}).values(), key=lambda item: (item.blogger, item.published_at, item.topic))
        write_category_table(output_dir / filename, category, rows)

    sync_selection_outputs(output_dir, audit, totals)
    totals["other_total"] = len(existing.get("其他类型", {}))

    legacy_manual_exists_after, legacy_manual_mtime_after = legacy_manual_table_state(output_dir)
    if legacy_manual_exists_before and legacy_manual_exists_after and legacy_manual_mtime_before != legacy_manual_mtime_after:
        raise SystemExit("保护失败：旧入口 00_手动输入选题表.md 不应被直接改写")

    audit_path = write_audit(audit_dir, audit, totals, output_dir)
    print({
        "inserted": totals["inserted"],
        "duplicates": totals["duplicates"],
        "moved": totals["moved"],
        "other": totals["other"],
        "audit_path": str(audit_path),
    })
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
